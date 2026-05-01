from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Q, Sum, Count
from django.http import HttpResponse, FileResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST
from .models import (
    Passenger, Group, GroupMembership, Document,
    Booking, Payment, Supplier, SupplierPayment, SERVICE_CHOICES,
    BookingHotel, BookingFlight, BookingTransport, BookingPassenger, Hotel, Airline,
)
from .forms import (
    PassengerForm, GroupForm, GroupMembershipForm, DocumentForm,
    BookingForm, PaymentForm, SupplierForm, SupplierPaymentForm,
    BookingHotelForm, BookingFlightForm, BookingTransportForm, BookingPassengerForm,
    HotelForm, AirlineForm,
)


def _save_with_audit(form, user, commit=True):
    """Save a ModelForm and stamp created_by/updated_by from the request user."""
    obj = form.save(commit=False)
    if not obj.pk and hasattr(obj, "created_by_id") and not obj.created_by_id:
        obj.created_by = user
    if hasattr(obj, "updated_by_id"):
        obj.updated_by = user
    if commit:
        obj.save()
    return obj


def _set_active(request, model, pk, active, redirect_name, perm):
    if not request.user.has_perm(perm):
        messages.error(request, "Permission denied.")
        return redirect(redirect_name)
    obj = get_object_or_404(model, pk=pk)
    obj.is_active = active
    if hasattr(obj, "updated_by_id"):
        obj.updated_by = request.user
    obj.save(update_fields=["is_active", "updated_by", "updated_at"] if hasattr(obj, "updated_by_id") else ["is_active", "updated_at"])
    messages.success(request, "Restored." if active else "Archived. Visit the Archived list to restore.")
    return redirect(redirect_name)


# ---------- Dashboard ----------

@login_required
def dashboard(request):
    today = timezone.now().date()
    soon_passport = today + timedelta(days=180)
    soon_visa = today + timedelta(days=30)

    expiring_passports = Passenger.objects.filter(
        passport_expiry__isnull=False, passport_expiry__lte=soon_passport
    ).order_by("passport_expiry")[:10]

    expiring_visas = Passenger.objects.filter(
        visa_expiry__isnull=False, visa_expiry__lte=soon_visa
    ).order_by("visa_expiry")[:10]

    today_bookings = Booking.objects.filter(created_at__date=today).count()
    pending_payments = Booking.objects.exclude(status="CANCELLED")
    total_billed = pending_payments.aggregate(s=Sum("package_cost"))["s"] or Decimal("0")
    total_received = Payment.objects.filter(booking__in=pending_payments).aggregate(s=Sum("amount"))["s"] or Decimal("0")
    outstanding = total_billed - total_received

    upcoming_departures = Group.objects.filter(
        departure_date__isnull=False, departure_date__gte=today
    ).order_by("departure_date")[:5]

    by_service = (
        Booking.objects.values("service_type")
        .annotate(c=Count("id"))
        .order_by("-c")
    )

    ctx = {
        "expiring_passports": expiring_passports,
        "expiring_visas": expiring_visas,
        "today_bookings": today_bookings,
        "total_billed": total_billed,
        "total_received": total_received,
        "outstanding": outstanding,
        "upcoming_departures": upcoming_departures,
        "by_service": by_service,
        "recent_bookings": Booking.objects.select_related("passenger")[:8],
        "passenger_count": Passenger.objects.count(),
        "group_count": Group.objects.count(),
    }
    return render(request, "crm/dashboard.html", ctx)


# ---------- Passengers ----------

@login_required
def passenger_list(request):
    q = (request.GET.get("q") or "").strip()
    show_archived = request.GET.get("archived") == "1"
    qs = Passenger.objects.filter(is_active=not show_archived)
    if q:
        qs = qs.filter(
            Q(full_name__icontains=q)
            | Q(passport_number__icontains=q)
            | Q(cnic__icontains=q)
            | Q(mobile__icontains=q)
            | Q(email__icontains=q)
        )
    return render(request, "crm/passenger_list.html", {"passengers": qs[:500], "q": q, "total": qs.count(), "show_archived": show_archived})


@login_required
@permission_required("crm.add_passenger", raise_exception=True)
def passenger_create(request):
    form = PassengerForm(request.POST or None)
    if form.is_valid():
        p = _save_with_audit(form, request.user)
        messages.success(request, "Passenger created.")
        return redirect(p.get_absolute_url())
    return render(request, "crm/passenger_form.html", {"form": form, "title": "New Passenger"})


@login_required
def passenger_detail(request, pk):
    p = get_object_or_404(Passenger, pk=pk)
    doc_form = DocumentForm()
    return render(request, "crm/passenger_detail.html", {
        "p": p, "doc_form": doc_form,
        "bookings": p.bookings.all(),
    })


@login_required
@permission_required("crm.change_passenger", raise_exception=True)
def passenger_edit(request, pk):
    p = get_object_or_404(Passenger, pk=pk)
    form = PassengerForm(request.POST or None, instance=p)
    if form.is_valid():
        _save_with_audit(form, request.user)
        messages.success(request, "Passenger updated.")
        return redirect(p.get_absolute_url())
    return render(request, "crm/passenger_form.html", {"form": form, "title": "Edit Passenger"})


@login_required
@permission_required("crm.delete_passenger", raise_exception=True)
@require_POST
def passenger_delete(request, pk):
    return _set_active(request, Passenger, pk, False, "passenger_list", "crm.delete_passenger")


@login_required
@permission_required("crm.change_passenger", raise_exception=True)
@require_POST
def passenger_restore(request, pk):
    return _set_active(request, Passenger, pk, True, "passenger_list", "crm.change_passenger")


@login_required
@permission_required("crm.add_document", raise_exception=True)
@require_POST
def document_upload(request, pk):
    p = get_object_or_404(Passenger, pk=pk)
    form = DocumentForm(request.POST, request.FILES)
    if form.is_valid():
        doc = form.save(commit=False)
        doc.passenger = p
        doc.uploaded_by = request.user
        doc.save()
        messages.success(request, "Document uploaded.")
    else:
        messages.error(request, "Upload failed.")
    return redirect(p.get_absolute_url())


@login_required
@permission_required("crm.delete_document", raise_exception=True)
@require_POST
def document_delete(request, pk):
    d = get_object_or_404(Document, pk=pk)
    pid = d.passenger_id
    d.file.delete(save=False)
    d.delete()
    return redirect("passenger_detail", pk=pid)


# ---------- Groups ----------

@login_required
def group_list(request):
    q = (request.GET.get("q") or "").strip()
    qs = Group.objects.all()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(destination__icontains=q))
    return render(request, "crm/group_list.html", {"groups": qs, "q": q})


@login_required
@permission_required("crm.add_group", raise_exception=True)
def group_create(request):
    form = GroupForm(request.POST or None)
    if form.is_valid():
        g = form.save()
        messages.success(request, "Group created.")
        return redirect(g.get_absolute_url())
    return render(request, "crm/group_form.html", {"form": form, "title": "New Group"})


@login_required
def group_detail(request, pk):
    g = get_object_or_404(Group, pk=pk)
    add_form = GroupMembershipForm()
    return render(request, "crm/group_detail.html", {"g": g, "add_form": add_form})


@login_required
@permission_required("crm.change_group", raise_exception=True)
def group_edit(request, pk):
    g = get_object_or_404(Group, pk=pk)
    form = GroupForm(request.POST or None, instance=g)
    if form.is_valid():
        form.save()
        return redirect(g.get_absolute_url())
    return render(request, "crm/group_form.html", {"form": form, "title": "Edit Group"})


@login_required
@permission_required("crm.delete_group", raise_exception=True)
@require_POST
def group_delete(request, pk):
    get_object_or_404(Group, pk=pk).delete()
    return redirect("group_list")


@login_required
@permission_required("crm.add_groupmembership", raise_exception=True)
@require_POST
def group_add_member(request, pk):
    g = get_object_or_404(Group, pk=pk)
    form = GroupMembershipForm(request.POST)
    if form.is_valid():
        m = form.save(commit=False)
        m.group = g
        try:
            m.save()
            messages.success(request, "Member added.")
        except Exception:
            messages.error(request, "Already a member.")
    return redirect(g.get_absolute_url())


@login_required
@permission_required("crm.delete_groupmembership", raise_exception=True)
@require_POST
def group_remove_member(request, pk, passenger_id):
    GroupMembership.objects.filter(group_id=pk, passenger_id=passenger_id).delete()
    return redirect("group_detail", pk=pk)


# ---------- Bookings ----------

@login_required
def booking_list(request):
    q = (request.GET.get("q") or "").strip()
    status = request.GET.get("status") or ""
    service = request.GET.get("service") or ""
    show_archived = request.GET.get("archived") == "1"
    qs = Booking.objects.select_related("passenger", "group").filter(is_active=not show_archived)
    if q:
        qs = qs.filter(
            Q(reference__icontains=q)
            | Q(passenger__full_name__icontains=q)
            | Q(passenger__passport_number__icontains=q)
        )
    if status:
        qs = qs.filter(status=status)
    if service:
        qs = qs.filter(service_type=service)
    return render(request, "crm/booking_list.html", {
        "bookings": qs[:500], "q": q, "status": status, "service": service,
        "service_choices": SERVICE_CHOICES, "show_archived": show_archived,
    })


@login_required
@login_required
@permission_required("crm.add_booking", raise_exception=True)
def booking_create(request):
    form = BookingForm(request.POST or None)
    if form.is_valid():
        b = _save_with_audit(form, request.user)
        messages.success(request, f"Booking {b.reference} created.")
        return redirect(b.get_absolute_url())
    return render(request, "crm/booking_form.html", {"form": form, "title": "New Booking"})


@login_required
def booking_detail(request, pk):
    b = get_object_or_404(Booking.objects.select_related("passenger", "group"), pk=pk)
    pay_form = PaymentForm()
    return render(request, "crm/booking_detail.html", {
        "b": b,
        "pay_form": pay_form,
        "hotel_form": BookingHotelForm(),
        "flight_form": BookingFlightForm(),
        "transport_form": BookingTransportForm(),
        "pax_form": BookingPassengerForm(),
    })


@login_required
@permission_required("crm.change_booking", raise_exception=True)
def booking_edit(request, pk):
    b = get_object_or_404(Booking, pk=pk)
    form = BookingForm(request.POST or None, instance=b)
    if form.is_valid():
        _save_with_audit(form, request.user)
        return redirect(b.get_absolute_url())
    return render(request, "crm/booking_form.html", {"form": form, "title": "Edit Booking"})


@login_required
@permission_required("crm.delete_booking", raise_exception=True)
@require_POST
def booking_delete(request, pk):
    return _set_active(request, Booking, pk, False, "booking_list", "crm.delete_booking")


@login_required
@permission_required("crm.change_booking", raise_exception=True)
@require_POST
def booking_restore(request, pk):
    return _set_active(request, Booking, pk, True, "booking_list", "crm.change_booking")


@login_required
@permission_required("crm.add_payment", raise_exception=True)
@require_POST
def payment_create(request, pk):
    b = get_object_or_404(Booking, pk=pk)
    form = PaymentForm(request.POST)
    if form.is_valid():
        p = form.save(commit=False)
        p.booking = b
        p.received_by = request.user
        p.save()
        messages.success(request, "Payment recorded.")
    else:
        messages.error(request, "Invalid payment.")
    return redirect(b.get_absolute_url())


# ---------- Suppliers ----------

@login_required
def supplier_list(request):
    return render(request, "crm/supplier_list.html", {"suppliers": Supplier.objects.all()})


@login_required
@permission_required("crm.add_supplier", raise_exception=True)
def supplier_create(request):
    form = SupplierForm(request.POST or None)
    if form.is_valid():
        s = form.save()
        return redirect("supplier_detail", pk=s.pk)
    return render(request, "crm/supplier_form.html", {"form": form, "title": "New Supplier"})


@login_required
def supplier_detail(request, pk):
    s = get_object_or_404(Supplier, pk=pk)
    pay_form = SupplierPaymentForm(initial={"supplier": s})
    return render(request, "crm/supplier_detail.html", {"s": s, "pay_form": pay_form})


@login_required
@permission_required("crm.change_supplier", raise_exception=True)
def supplier_edit(request, pk):
    s = get_object_or_404(Supplier, pk=pk)
    form = SupplierForm(request.POST or None, instance=s)
    if form.is_valid():
        form.save()
        return redirect("supplier_detail", pk=s.pk)
    return render(request, "crm/supplier_form.html", {"form": form, "title": "Edit Supplier"})


@login_required
@permission_required("crm.add_supplierpayment", raise_exception=True)
@require_POST
def supplier_payment_create(request, pk):
    s = get_object_or_404(Supplier, pk=pk)
    form = SupplierPaymentForm(request.POST)
    if form.is_valid():
        p = form.save(commit=False)
        p.supplier = s
        p.save()
        messages.success(request, "Supplier payment recorded.")
    return redirect("supplier_detail", pk=s.pk)


# ---------- Reports / Exports ----------

def _excel_response(filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_passengers_excel(request):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Passengers"
    headers = ["Full Name", "Passport #", "CNIC", "Passport Expiry", "Visa Expiry", "Mobile", "Email"]
    ws.append(headers)
    for p in Passenger.objects.all():
        ws.append([
            p.full_name, p.passport_number, p.cnic,
            p.passport_expiry.isoformat() if p.passport_expiry else "",
            p.visa_expiry.isoformat() if p.visa_expiry else "",
            p.mobile, p.email,
        ])
    resp = _excel_response("passengers.xlsx")
    wb.save(resp)
    return resp


@login_required
def export_passengers_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Passenger List")
    styles = getSampleStyleSheet()
    elements = [Paragraph("Passenger List", styles["Title"]), Spacer(1, 12)]
    data = [["#", "Name", "Passport", "CNIC", "P. Expiry", "V. Expiry", "Mobile"]]
    for i, p in enumerate(Passenger.objects.all(), 1):
        data.append([
            i, p.full_name, p.passport_number, p.cnic,
            p.passport_expiry or "", p.visa_expiry or "", p.mobile,
        ])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename="passengers.pdf")


@login_required
def booking_invoice_pdf(request, pk):
    b = get_object_or_404(Booking, pk=pk)
    return _booking_pdf(request, b, kind="invoice")


@login_required
def payment_receipt_pdf(request, pk, payment_id):
    b = get_object_or_404(Booking, pk=pk)
    payment = get_object_or_404(Payment, pk=payment_id, booking=b)
    return _booking_pdf(request, b, kind="receipt", payment=payment)


def _booking_pdf(request, b, kind="invoice", payment=None):
    from django.conf import settings
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=kind.title())
    styles = getSampleStyleSheet()
    el = []

    el.append(Paragraph(f"<b>{settings.AGENCY_NAME}</b>", styles["Title"]))
    el.append(Paragraph(settings.AGENCY_ADDRESS, styles["Normal"]))
    el.append(Paragraph(f"Phone: {settings.AGENCY_PHONE} &nbsp;&nbsp; Email: {settings.AGENCY_EMAIL}", styles["Normal"]))
    el.append(Spacer(1, 12))

    title = "INVOICE" if kind == "invoice" else "PAYMENT RECEIPT"
    el.append(Paragraph(f"<b>{title}</b>", styles["Heading2"]))

    info = [
        ["Reference", b.reference, "Date", timezone.now().date().isoformat()],
        ["Customer", b.passenger.full_name, "Passport", b.passenger.passport_number or "-"],
        ["Service", b.get_service_type_display(), "Travel Date", str(b.travel_date or "-")],
    ]
    t = Table(info, colWidths=[80, 180, 80, 180])
    t.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 9)]))
    el.append(t)
    el.append(Spacer(1, 12))

    if kind == "invoice":
        rows = [
            ["Description", "Amount"],
            [b.description or b.get_service_type_display(), f"{b.package_cost:.2f}"],
            ["Discount", f"-{b.discount:.2f}"],
            ["Net Total", f"{b.net_total:.2f}"],
            ["Received", f"{b.total_received:.2f}"],
            ["Balance Due", f"{b.balance_due:.2f}"],
        ]
    else:
        rows = [
            ["Item", "Detail"],
            ["Amount Received", f"{payment.amount:.2f}"],
            ["Method", payment.get_method_display()],
            ["Date", str(payment.received_on)],
            ["Reference", payment.reference or "-"],
            ["Booking Net Total", f"{b.net_total:.2f}"],
            ["Balance After", f"{b.balance_due:.2f}"],
        ]

    t2 = Table(rows, colWidths=[300, 200])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    el.append(t2)
    el.append(Spacer(1, 24))
    el.append(Paragraph("Thank you for choosing us.", styles["Italic"]))

    doc.build(el)
    buf.seek(0)
    fname = f"{kind}-{b.reference}.pdf"
    return FileResponse(buf, as_attachment=True, filename=fname)


@login_required
def financial_report(request):
    bookings = Booking.objects.select_related("passenger").exclude(status="CANCELLED")
    total_billed = bookings.aggregate(s=Sum("package_cost"))["s"] or Decimal("0")
    total_disc = bookings.aggregate(s=Sum("discount"))["s"] or Decimal("0")
    total_received = Payment.objects.filter(booking__in=bookings).aggregate(s=Sum("amount"))["s"] or Decimal("0")
    supplier_paid = SupplierPayment.objects.aggregate(s=Sum("amount"))["s"] or Decimal("0")
    return render(request, "crm/financial_report.html", {
        "bookings": bookings[:200],
        "total_billed": total_billed,
        "total_disc": total_disc,
        "total_received": total_received,
        "outstanding": total_billed - total_disc - total_received,
        "supplier_paid": supplier_paid,
    })


@login_required
def export_financial_excel(request):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"
    ws.append(["Reference", "Customer", "Service", "Package", "Discount", "Net", "Received", "Balance", "Status"])
    for b in Booking.objects.select_related("passenger"):
        ws.append([
            b.reference, b.passenger.full_name, b.get_service_type_display(),
            float(b.package_cost), float(b.discount), float(b.net_total),
            float(b.total_received), float(b.balance_due), b.get_status_display(),
        ])
    ws2 = wb.create_sheet("Supplier Payments")
    ws2.append(["Date", "Supplier", "Amount", "Method", "Booking", "Reference"])
    for sp in SupplierPayment.objects.select_related("supplier", "booking"):
        ws2.append([
            sp.paid_on.isoformat(), sp.supplier.name, float(sp.amount),
            sp.get_method_display(), sp.booking.reference if sp.booking else "",
            sp.reference,
        ])
    resp = _excel_response("financial_report.xlsx")
    wb.save(resp)
    return resp


# ---------- Booking inline items (hotels / flights / transports / extra pax) ----------

def _add_inline_item(request, pk, FormClass, fk_attr, success_msg, perm):
    if not request.user.has_perm(perm):
        messages.error(request, "Permission denied.")
        return redirect("booking_detail", pk=pk)
    b = get_object_or_404(Booking, pk=pk)
    form = FormClass(request.POST)
    if form.is_valid():
        obj = form.save(commit=False)
        setattr(obj, fk_attr, b)
        try:
            obj.save()
            messages.success(request, success_msg)
        except Exception as e:
            messages.error(request, f"Could not save: {e}")
    else:
        messages.error(request, "Invalid input.")
    return redirect("booking_detail", pk=pk)


@login_required
@require_POST
def booking_hotel_add(request, pk):
    return _add_inline_item(request, pk, BookingHotelForm, "booking", "Hotel added.", "crm.change_booking")


@login_required
@require_POST
def booking_hotel_delete(request, pk, item_id):
    if not request.user.has_perm("crm.change_booking"):
        messages.error(request, "Permission denied.")
        return redirect("booking_detail", pk=pk)
    BookingHotel.objects.filter(pk=item_id, booking_id=pk).delete()
    return redirect("booking_detail", pk=pk)


@login_required
@require_POST
def booking_flight_add(request, pk):
    return _add_inline_item(request, pk, BookingFlightForm, "booking", "Flight added.", "crm.change_booking")


@login_required
@require_POST
def booking_flight_delete(request, pk, item_id):
    if not request.user.has_perm("crm.change_booking"):
        messages.error(request, "Permission denied.")
        return redirect("booking_detail", pk=pk)
    BookingFlight.objects.filter(pk=item_id, booking_id=pk).delete()
    return redirect("booking_detail", pk=pk)


@login_required
@require_POST
def booking_transport_add(request, pk):
    return _add_inline_item(request, pk, BookingTransportForm, "booking", "Transport added.", "crm.change_booking")


@login_required
@require_POST
def booking_transport_delete(request, pk, item_id):
    if not request.user.has_perm("crm.change_booking"):
        messages.error(request, "Permission denied.")
        return redirect("booking_detail", pk=pk)
    BookingTransport.objects.filter(pk=item_id, booking_id=pk).delete()
    return redirect("booking_detail", pk=pk)


@login_required
@require_POST
def booking_pax_add(request, pk):
    return _add_inline_item(request, pk, BookingPassengerForm, "booking", "Pilgrim added.", "crm.change_booking")


@login_required
@require_POST
def booking_pax_delete(request, pk, item_id):
    if not request.user.has_perm("crm.change_booking"):
        messages.error(request, "Permission denied.")
        return redirect("booking_detail", pk=pk)
    BookingPassenger.objects.filter(pk=item_id, booking_id=pk).delete()
    return redirect("booking_detail", pk=pk)


# ---------- Hotel Voucher PDF ----------

@login_required
def booking_voucher_pdf(request, pk):
    b = get_object_or_404(
        Booking.objects.select_related("passenger").prefetch_related(
            "hotels", "flights", "transports", "extra_pax__passenger"
        ),
        pk=pk,
    )
    return _voucher_pdf(b)


def _voucher_pdf(b):
    from django.conf import settings
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=f"Hotel Voucher {b.effective_voucher_no}",
    )

    styles = getSampleStyleSheet()
    h_center = ParagraphStyle("hc", parent=styles["Heading2"], alignment=TA_CENTER, spaceAfter=4)
    title_box = ParagraphStyle("tb", parent=styles["Heading2"], alignment=TA_CENTER, fontSize=14)
    small = ParagraphStyle("sm", parent=styles["Normal"], fontSize=8)

    el = []

    # Agency name top
    el.append(Paragraph(f"<b>{settings.AGENCY_NAME}</b>", h_center))
    el.append(Spacer(1, 4))

    # Hotel Voucher banner
    banner = Table([[Paragraph("<b>Hotel Voucher</b>", title_box)]], colWidths=[80 * mm])
    banner.hAlign = "CENTER"
    banner.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    el.append(banner)
    el.append(Spacer(1, 6))

    # Header info table (4 cols)
    fmt_d = lambda d: d.strftime("%d/%m/%y") if d else ""
    voucher_date = b.voucher_date or b.created_at.date()
    head_rows = [
        ["IATA:", settings.AGENCY_NAME, "Voucher No:", b.effective_voucher_no, "Branch:", b.branch or ""],
        ["Saudi Company:", b.saudi_company or "", "Date:", fmt_d(voucher_date), "", ""],
        ["Package:", b.package_label or b.get_service_type_display(), "Manual No:", b.manual_no or "", "Group #:", b.group_no or ""],
        ["Family Head:", b.passenger.full_name.upper(), "Total PAX:", str(b.total_pax), "Whatsapp:", b.whatsapp or ""],
    ]
    th = Table(head_rows, colWidths=[24 * mm, 50 * mm, 22 * mm, 36 * mm, 20 * mm, 34 * mm])
    th.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f1f1")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f1f1f1")),
        ("BACKGROUND", (4, 0), (4, -1), colors.HexColor("#f1f1f1")),
    ]))
    el.append(th)

    # Pilgrims Details
    el.append(_section_bar("Pilgrims Details"))
    pilgrim_rows = [["Mutamer Name", "Gender", "PPNO", "PAX", "Beds", "Visa Number", "PNR"]]
    for p in b.all_pilgrims:
        pax = p["passenger"]
        pilgrim_rows.append([
            pax.full_name.upper(),
            pax.get_gender_display(),
            pax.passport_number or "",
            p["pax_type_display"],
            "Yes" if p["bed"] else "No",
            p["visa_number"],
            p["pnr"],
        ])
    pt = Table(pilgrim_rows, colWidths=[55 * mm, 18 * mm, 25 * mm, 15 * mm, 13 * mm, 30 * mm, 30 * mm])
    pt.setStyle(_table_style(header=True))
    el.append(pt)

    # Totals row
    summary = Table(
        [[f"Adults: {b.adults_count}", f"Childs: {b.children_count}",
          f"Infants: {b.infants_count}", f"Total PAX: {b.total_pax}",
          f"Total BEDS: {b.total_beds}"]],
        colWidths=[30 * mm, 25 * mm, 27 * mm, 30 * mm, 30 * mm],
    )
    summary.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    el.append(summary)

    # Accommodation
    el.append(_section_bar("Accommodation Details:"))
    hotel_rows = [["Hotel Name", "Confirm No", "City", "Room", "Meal Plan", "Check In", "Checkout", "Nights"]]
    total_nights = 0
    for h in b.hotels.all():
        room_cell = h.room_label
        if h.room_notes:
            room_cell += f"\n({h.room_notes})"
        hotel_rows.append([
            h.hotel.name, h.confirm_no, h.hotel.city, room_cell, h.meal_plan,
            fmt_d(h.check_in), fmt_d(h.check_out), str(h.nights),
        ])
        total_nights += h.nights
    if len(hotel_rows) == 1:
        hotel_rows.append(["—", "", "", "", "", "", "", ""])
    ht = Table(hotel_rows, colWidths=[45 * mm, 22 * mm, 20 * mm, 28 * mm, 18 * mm, 18 * mm, 18 * mm, 14 * mm])
    ht.setStyle(_table_style(header=True))
    el.append(ht)
    tn = Table([["", "", "", "", "", "", "Total Nights:", str(total_nights)]],
               colWidths=[45 * mm, 22 * mm, 20 * mm, 28 * mm, 18 * mm, 18 * mm, 18 * mm, 14 * mm])
    tn.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (6, 0), (7, 0), 0.5, colors.black),
        ("FONTNAME", (6, 0), (7, 0), "Helvetica-Bold"),
        ("ALIGN", (7, 0), (7, 0), "CENTER"),
    ]))
    el.append(tn)

    # Transport / Services
    el.append(_section_bar("Transport / Services"))
    tr_rows = [["Name", "Mode", "Type / Details", "BRN"]]
    for t in b.transports.all():
        tr_rows.append([t.name, t.get_transport_mode_display(), t.transport_type, t.brn])
    if len(tr_rows) == 1:
        tr_rows.append(["—", "", "", ""])
    tr = Table(tr_rows, colWidths=[55 * mm, 35 * mm, 60 * mm, 33 * mm])
    tr.setStyle(_table_style(header=True))
    el.append(tr)

    # Flights — split into two side-by-side mini-tables
    out_flights = [f for f in b.flights.all() if f.direction == "OUT"]
    in_flights = [f for f in b.flights.all() if f.direction == "IN"]

    def _flight_table(title, flights):
        rows = [[Paragraph(f"<b>{title}</b>", small)]]
        rows.append(["Airline", "Flight", "Sector", "Departure", "Arrival"])
        for f in flights:
            rows.append([
                f.airline.name if f.airline else "",
                f.flight_no, f.sector,
                f.departure.strftime("%d-%b %H:%M") if f.departure else "",
                f.arrival.strftime("%d-%b %H:%M") if f.arrival else "",
            ])
        if len(rows) == 2:
            rows.append(["—", "", "", "", ""])
        t = Table(rows, colWidths=[18 * mm, 14 * mm, 16 * mm, 19 * mm, 19 * mm])
        t.setStyle(TableStyle([
            ("SPAN", (0, 0), (-1, 0)),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9e9e9")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f1f1f1")),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    flights_tbl = Table(
        [[_flight_table("Departure from Pakistan to KSA", out_flights),
          _flight_table("Departure from KSA to Pakistan", in_flights)]],
        colWidths=[88 * mm, 88 * mm],
    )
    flights_tbl.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    el.append(Spacer(1, 4))
    el.append(flights_tbl)

    # Footer note
    if b.voucher_note:
        el.append(Spacer(1, 6))
        el.append(Paragraph(f"<b>Note:</b> {b.voucher_note}", small))

    el.append(Spacer(1, 4))
    el.append(Paragraph(
        f"Phone: {settings.AGENCY_PHONE} &nbsp;&nbsp; Email: {settings.AGENCY_EMAIL} &nbsp;&nbsp; {settings.AGENCY_ADDRESS}",
        small,
    ))

    doc.build(el)
    buf.seek(0)
    return FileResponse(buf, as_attachment=False, filename=f"voucher-{b.effective_voucher_no}.pdf")


def _section_bar(text):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    t = Table([[text]], colWidths=[186 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#dcdcdc")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


def _table_style(header=False):
    from reportlab.platypus import TableStyle
    from reportlab.lib import colors
    s = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]
    if header:
        s += [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f1f1")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]
    return TableStyle(s)


# ---------- Hotels ----------

@login_required
def hotel_list(request):
    q = (request.GET.get("q") or "").strip()
    show_archived = request.GET.get("archived") == "1"
    qs = Hotel.objects.filter(is_active=not show_archived)
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(city__icontains=q))
    form = HotelForm()
    return render(request, "crm/hotel_list.html", {"hotels": qs, "q": q, "form": form, "show_archived": show_archived})


@login_required
@permission_required("crm.add_hotel", raise_exception=True)
@require_POST
def hotel_create(request):
    form = HotelForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, "Hotel added.")
    else:
        messages.error(request, "Could not add hotel: " + " | ".join(f"{k}: {','.join(v)}" for k, v in form.errors.items()))
    return redirect("hotel_list")


@login_required
@permission_required("crm.change_hotel", raise_exception=True)
def hotel_edit(request, pk):
    h = get_object_or_404(Hotel, pk=pk)
    form = HotelForm(request.POST or None, instance=h)
    if form.is_valid():
        form.save()
        messages.success(request, "Hotel updated.")
        return redirect("hotel_list")
    return render(request, "crm/hotel_form.html", {"form": form, "h": h, "title": "Edit Hotel"})


@login_required
@permission_required("crm.delete_hotel", raise_exception=True)
@require_POST
def hotel_delete(request, pk):
    return _set_active(request, Hotel, pk, False, "hotel_list", "crm.delete_hotel")


@login_required
@permission_required("crm.change_hotel", raise_exception=True)
@require_POST
def hotel_restore(request, pk):
    return _set_active(request, Hotel, pk, True, "hotel_list", "crm.change_hotel")


# ---------- Airlines ----------

@login_required
def airline_list(request):
    q = (request.GET.get("q") or "").strip()
    show_archived = request.GET.get("archived") == "1"
    qs = Airline.objects.filter(is_active=not show_archived)
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q) | Q(country__icontains=q))
    form = AirlineForm()
    return render(request, "crm/airline_list.html", {"airlines": qs, "q": q, "form": form, "show_archived": show_archived})


@login_required
@permission_required("crm.add_airline", raise_exception=True)
@require_POST
def airline_create(request):
    form = AirlineForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, "Airline added.")
    else:
        messages.error(request, "Could not add airline: " + " | ".join(f"{k}: {','.join(v)}" for k, v in form.errors.items()))
    return redirect("airline_list")


@login_required
@permission_required("crm.change_airline", raise_exception=True)
def airline_edit(request, pk):
    a = get_object_or_404(Airline, pk=pk)
    form = AirlineForm(request.POST or None, instance=a)
    if form.is_valid():
        form.save()
        messages.success(request, "Airline updated.")
        return redirect("airline_list")
    return render(request, "crm/airline_form.html", {"form": form, "a": a, "title": "Edit Airline"})


@login_required
@permission_required("crm.delete_airline", raise_exception=True)
@require_POST
def airline_delete(request, pk):
    return _set_active(request, Airline, pk, False, "airline_list", "crm.delete_airline")


@login_required
@permission_required("crm.change_airline", raise_exception=True)
@require_POST
def airline_restore(request, pk):
    return _set_active(request, Airline, pk, True, "airline_list", "crm.change_airline")
